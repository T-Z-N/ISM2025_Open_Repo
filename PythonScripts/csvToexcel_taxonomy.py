import pandas as pd
from openpyxl.styles import Alignment

# CSV file path
csv_file_path = 'taxonomy_query_results.csv'
# Excel file path to create
excel_file_path = 'taxonomy_hierarchy.xlsx'

try:
    # Read the CSV file
    df = pd.read_csv(csv_file_path)
    
    # Replace NaN values with empty strings for better visual appearance
    df.fillna('', inplace=True)
    
    # Create Excel writer
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Taxonomy', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Taxonomy']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Merge cells for superclass column where values are the same
        if not df.empty:
            current_superclass_uri = None
            merge_start_row = 2  # First data row in Excel (after header)
            
            for idx in range(len(df)):
                # Current Excel row number (1-based)
                excel_current_row = idx + 2
                
                # If this is the first row or superclass changed
                if current_superclass_uri is None or df.loc[idx, 'superclass'] != current_superclass_uri:
                    # Merge previous superclass block if it spans multiple rows
                    if current_superclass_uri is not None and excel_current_row - 1 > merge_start_row:
                        worksheet.merge_cells(start_row=merge_start_row, start_column=1,
                                              end_row=excel_current_row - 1, end_column=1)
                        # Center align the merged cell
                        merged_cell = worksheet.cell(row=merge_start_row, column=1)
                        merged_cell.alignment = Alignment(vertical='center')
                    
                    # Set new superclass block start
                    current_superclass_uri = df.loc[idx, 'superclass']
                    merge_start_row = excel_current_row
            
            # Merge the last block after the loop
            if current_superclass_uri is not None and len(df) + 1 > merge_start_row:
                worksheet.merge_cells(start_row=merge_start_row, start_column=1,
                                      end_row=len(df) + 1, end_column=1)
                # Center align the merged cell
                merged_cell = worksheet.cell(row=merge_start_row, column=1)
                merged_cell.alignment = Alignment(vertical='center')
    
    print(f"Excel file '{excel_file_path}' has been successfully created.")
    print(f"The file contains {len(df)} taxonomy relationships.")

except FileNotFoundError:
    print(f"Error: File '{csv_file_path}' not found.")
except Exception as e:
    print(f"An error occurred: {e}")